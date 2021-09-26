import urllib.parse
import logging
import xlrd
import json
from robot.libraries.BuiltIn import BuiltIn
# import pyodbc
from dateutil.relativedelta import relativedelta
import datetime
import  openpyxl
import  time


def set_query_parameter(url, param):
    # logging.info(param)
    # BuiltIn().set_test_message(param)
    logging.warn(param)
    if  (param is None) or (param == {}):
        return url
    else:
        return url+'?'+urllib.parse.urlencode(param)


def make_json_from_data(column_names, row_data):
    row_list = []
    for item in row_data:
        json_obj = {}
        for i in range(0, column_names.__len__()):
            json_obj[column_names[i]] = to_json(item[i])
        row_list.append(json_obj)
    return row_list

def to_json(myjson):
    try:
        json_object = json.loads(str(myjson))
    except ValueError as e:
        return myjson
    return json_object


def xls_to_dict(workbook_url):
    workbook_dict = {}
    book = xlrd.open_workbook(workbook_url)
    sheets = book.sheets()
    for sheet in sheets:
        if sheet.name == 'PortHoles & Discrete Appurtenan':
            continue
        workbook_dict[sheet.name] = {}
        columns = sheet.row_values(0)
        rows = []
        for row_index in range(1, sheet.nrows):
            row = sheet.row_values(row_index)
            rows.append(row)
        sheet_data = make_json_from_data(columns, rows)
        workbook_dict[sheet.name] = sheet_data
    return workbook_dict
    
def get_data_test_case(workbook_url, sheetName ,testcaseName):
    json_object = xls_to_dict(workbook_url)
    for dict in json_object[sheetName]:
        if dict['TC No.'] == testcaseName:
            return dict


def select_sql(connection_string,query):
    cnxn = pyodbc.connect(connection_string)
    cursor = cnxn.cursor()
    cursor.execute(query) 
    rows = cursor.fetchone() 
    return rows
 
def insert_sq_sql(connection_string, **sq_data):
    cnxn = pyodbc.connect(connection_string)
    cursor = cnxn.cursor()
    cursor.execute('EXEC [dbo].[proc_create_newcombination] ' +sq_data['sq_ty']+' , '+ sq_data['chnl_cd']+' , '+ sq_data['rgn_cd']+' , '+ sq_data['sls_grp_cd']+' , '+ sq_data['cust_cd']+' , '+ sq_data['mat_cd']+' , '+ sq_data['mth_cd']+' , '+ sq_data['init_vol'])
    cnxn.commit()

def reset_sq_sql(connection_string):
    cnxn = pyodbc.connect(connection_string)
    cursor = cnxn.cursor()
    cursor.execute('EXEC [dbo].[proc_reset_init_SQ]')
    cnxn.commit()
 
def reset_month_end_rollSQ_onhnd(connection_string):
    cnxn = pyodbc.connect(connection_string)
    cursor = cnxn.cursor()
    cursor.execute('EXEC [dbo].[proc_reset_month_end_rollSQ_onhnd]')
    cnxn.commit()
 
def reset_month_end_rollSQ_remain(connection_string):
    cnxn = pyodbc.connect(connection_string)
    cursor = cnxn.cursor()
    cursor.execute('EXEC [dbo].[proc_reset_month_end_rollSQ_remain]')
    cnxn.commit()

def reset_SQ_mat_ctrl(connection_string):
    cnxn = pyodbc.connect(connection_string)
    cursor = cnxn.cursor()
    cursor.execute('EXEC [dbo].[proc_reset_SQ_mat_ctrl]')
    cnxn.commit()

def reset_rsrv(connection_string):
    cnxn = pyodbc.connect(connection_string)
    cursor = cnxn.cursor()
    cursor.execute('EXEC [dbo].[proc_reset_rsrv]')
    cnxn.commit()

def reset_atp(connection_string):
    cnxn = pyodbc.connect(connection_string)
    cursor = cnxn.cursor()
    cursor.execute('EXEC [dbo].[proc_reset_atp]')
    cnxn.commit()
   
def reset_on_hnd(connection_string):
    cnxn = pyodbc.connect(connection_string)
    cursor = cnxn.cursor()
    cursor.execute('EXEC [dbo].[proc_reset_on_hnd]')
    cnxn.commit()



def get_month(start_at=0, number_month=1, separator="/"):              #แก้ไข date ให้ไม่ต้องอ่านใน datatest
    if (number_month < 1):
        number_month = 1
    return_date = "" #2020/07
    for index in range(0, number_month):
        add_month = start_at + index
        get_date = datetime.date.today() + relativedelta(months=add_month)
        get_year = str(get_date.year)
        get_month = str(get_date.month) if get_date.month > 9 else "0" + str(get_date.month)
        data_date = get_year + separator + get_month #2020/08
        if (index != 0):
            data_date = "," + data_date #,2020/08
        return_date = return_date + data_date #2020/07,2020/08
    return return_date #2020/07,2020/08


def get_today():            #load date Json
    get_date = datetime.date.today()
    get_year = str(get_date.year)
    get_month = str(get_date.month) if get_date.month > 9 else "0" + str(get_date.month)
    get_day = str(get_date.day) if get_date.day > 9 else "0" + str(get_date.day)
    data_date = get_year + "/" + get_month + "/" + get_day
    return data_date

# print(get_today())

# def get_today_timestamp():     #timestamp Json (epoch time)
#     ts = datetime.datetime.now().timestamp()
#     return int(ts)

def get_today_timestamp(addstring=""):                      #Normal Timestamp Json (Today)
    addstring = str(addstring)
    ts = datetime.datetime.now().timestamp()
    ts = int(ts)
    return str(ts) + addstring

def set_data_excel(workbook_url, sheetname, column_data, new_data):             #Update date on Excel file
    xfile = openpyxl.load_workbook(workbook_url)
    sheet = xfile[sheetname]
    sheet[column_data] = new_data
    xfile.save(workbook_url)

def get_data_excel(workbook_url, sheetname, column_data):               #Update date on Excel file
    xfile = openpyxl.load_workbook(workbook_url)
    sheet = xfile[sheetname]
    return sheet[column_data].value
# set_data_excel("Resource/datatest-2.xlsx", "Month_End", "A1", get_month())
# get_data_excel("Resource/datatest-2.xlsx", "Month_End", "A1", get_month())
# print( get_data_excel("Resource/Bidding_Price_Table_Template_PE.xlsx", "BIDDING_Price_Table_Template", "G2"))

def get_week_cycle(previous_day = -7):                                  #Return Vol Bucket
    get_date = datetime.datetime.today() + relativedelta(days=previous_day)

    day_of_week = get_date.weekday()

    tuesday_of_week = 1 if day_of_week > 0 else -6
    monday_of_week = 7 if day_of_week > 0 else 0

    cal_day_first = tuesday_of_week - day_of_week
    cal_day_last = monday_of_week - day_of_week
    
    first_date = get_date + relativedelta(days=cal_day_first)
    last_date = get_date + relativedelta(days=cal_day_last)

    get_year = str(first_date.year)
    get_month = str(first_date.month) if first_date.month > 9 else "0" + str(first_date.month)
    get_day_first = str(first_date.day) if first_date.day > 9 else "0" + str(first_date.day)
    get_day_last = str(last_date.day) if last_date.day > 9 else "0" + str(last_date.day)

    data_date = get_year + get_month + "_" + get_day_first + "-" + get_day_last
    return data_date

def get_previous_day_timestamp(previous_day = -7):                          #Return Vol Bucket                
    get_date = datetime.datetime.now() + relativedelta(days=previous_day)
    timestamp = time.mktime(get_date.timetuple()) + get_date.microsecond/1e6
    timestamp = int(timestamp)
    timestamp = str(timestamp) + "000"
    return timestamp

# print(get_previous_day_timestamp())
    
def get_day_frozen(start_at=0, separator="/"):                  #สำหรับเคส Frozen โดยเฉพาะ เพราะเรียงตำแหน่งไม่เหมือนชาวบ้าน
    get_date = datetime.date.today() + relativedelta(days=start_at)
    get_year = str(get_date.year)
    get_month = str(get_date.month) if get_date.month > 9 else "0" + str(get_date.month)
    get_day = str(get_date.day) if get_date.day > 9 else "0" + str(get_date.day)
    data_date = get_day + separator + get_month + separator + get_year
    return data_date
# print(get_day_frozen())

def get_day(start_at=0, separator="/"):            #Today ใช้ได้เกือบทุกเคสที่มีการวางแบบนี้ 00/00/0000
    get_date = datetime.date.today() + relativedelta(days=start_at)
    get_year = str(get_date.year)
    get_month = str(get_date.month) if get_date.month > 9 else "0" + str(get_date.month)
    get_day = str(get_date.day) if get_date.day > 9 else "0" + str(get_date.day)
    data_date = get_year + separator + get_month + separator + get_day
    return data_date
# print(get_day_frozen())

def get_day_timestamp_frozen(previous_day = 1):         #สำหรับเคส Frozen เพื่อ Json PAM Check ยิงล่วงหน้า 1 วัน
    get_date = datetime.datetime.now() + relativedelta(days=previous_day)
    timestamp = time.mktime(get_date.timetuple()) + get_date.microsecond/1e6
    timestamp = int(timestamp)
    return timestamp
    
# print(get_day_timestamp_frozen())

def get_day2():            #ใช้สำหรับคลิ๊ก reserve ในหน้า Reservation แบบคลิ๊กจอง ณ วันที่ที่อัพโหลด ATP 
    get_date = datetime.date.today()
    get_day = str(get_date.day)
    return get_day

def count_row(self, sheet_name=None):
        sheet = self.get_sheet(sheet_name)
        cell = sheet.max_row
        return cell
def count_column(self, sheet_name=None):
        sheet = self.get_sheet(sheet_name)
        cell = sheet.max_column
        return cell  
  
def read_excel_column(self, col_num, row_offset=0, max_num=0,
                          sheet_name=None):
        # type: (int, int, int, str)->List[Any]
        col_num = int(col_num)
        row_offset = int(row_offset)
        max_num = int(max_num)
        sheet = self.get_sheet(sheet_name)
        row_iter = sheet.iter_rows(min_col=col_num, max_col=col_num,
                                   row_offset=row_offset,
                                   max_row=max_num)  # type: Iterator[Tuple[Cell]]
        return [row[0].value for row in row_iter]